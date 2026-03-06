"""
GSTR-3B Section-Aware Table Extractor (V3)

Extracts tables from GSTR-3B PDFs and maps them to their
specific section numbers (3.1, 3.1.1, 3.1.2, 4, 5, 5.1, 6).
Also extracts metadata: GSTIN, Filing Period, Date of ARN.
"""

import re
import logging
import pdfplumber
import pandas as pd
from typing import Dict, Any, List, Optional
from pathlib import Path

logger = logging.getLogger(__name__)

# Try tabula, but it's optional — pdfplumber is fallback
try:
    import tabula
    HAS_TABULA = True
except ImportError:
    HAS_TABULA = False

# ──────────────────────────────────────────────────────────────────────
# Section definitions
# ──────────────────────────────────────────────────────────────────────
GSTR3B_SECTIONS = {
    "3.1": {
        "title": "3.1 - Outward & Inward Supplies",
        "keywords": [
            "outward taxable supplies",
            "zero rated",
            "nil rated",
            "exempted",
            "inward supplies.*reverse charge",
            "non-gst outward",
        ],
    },
    "3.1.1": {
        "title": "3.1.1 - E-Commerce Supplies",
        "keywords": [
            "electronic commerce operator",
            "e-commerce",
            r"9\(5\)",
            "u/s 9",
        ],
    },
    "3.1.2": {
        "title": "3.1.2 - Inter-State Supplies",
        "keywords": [
            "unregistered persons",
            "composition taxable",
            "uin holders",
        ],
    },
    "4": {
        "title": "4 - Eligible ITC",
        "keywords": [
            "itc available",
            "import of goods",
            "import of services",
            "inward supplies.*isd",
            "all other itc",
            "itc reversed",
            "net itc",
            "ineligible itc",
        ],
    },
    "5": {
        "title": "5 - Exempt, Nil, Non-GST Supplies",
        "keywords": [
            "composition scheme.*exempt",
            "non gst supply",
            "inter.*state supplies",
            "intra.*state supplies",
        ],
    },
    "5.1": {
        "title": "5.1 - Interest & Late Fee",
        "keywords": [
            "system computed.*interest",
            "interest paid",
            "late fee",
        ],
    },
    "6": {
        "title": "6 - Tax Payment",
        "keywords": [
            "tax payable",
            "paid in cash",
            "paid through itc",
            "reverse charge",
            "net tax payable",
            "other than reverse charge",
        ],
    },
}


def clean_cell(value) -> str:
    """Clean a cell value: fix multiline artifacts, strip whitespace."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    s = str(value).strip()
    if s.lower() in ("none", "nan", "nat"):
        return ""
    # Fix tabula multiline number artifacts like "499509.0\n0" → "499509.00"
    s = re.sub(r'(\d+\.?\d*)\s*\n\s*(\d+)', r'\1\2', s)
    # Remove leading artifacts like "E\n", "F\n", "I\n"
    s = re.sub(r'^[A-Z]\s*\n\s*', '', s)
    # Replace remaining newlines with spaces
    s = s.replace('\n', ' ').replace('\r', '')
    # Collapse multiple spaces
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def _is_numeric_header(col_name: str) -> bool:
    """Check if a column name looks like a data value (not a real header)."""
    s = str(col_name).strip()
    if not s or s.lower() in ("none", "nan", ""):
        return True
    # Pure numbers like "0.00", "123456.78"
    try:
        float(s.replace(",", ""))
        return True
    except ValueError:
        pass
    # Unnamed columns from pandas
    if s.startswith("Unnamed:"):
        return True
    return False


def _fix_tabula_headers(df: pd.DataFrame) -> pd.DataFrame:
    """
    Fix tabula's tendency to put data values in column headers.
    If columns look like data (numbers, None, Unnamed), push them
    back as a data row and generate generic headers.
    """
    cols = list(df.columns)
    # Count how many columns look like data values
    bad_cols = sum(1 for c in cols if _is_numeric_header(c))

    if bad_cols > len(cols) * 0.4:
        # Most columns are data values — push header row back as data
        header_as_row = {f"Col_{i}": str(c) for i, c in enumerate(cols)}
        new_cols = [f"Col_{i}" for i in range(len(cols))]
        df.columns = new_cols
        # Prepend the old header as first data row
        header_df = pd.DataFrame([header_as_row])
        df = pd.concat([header_df, df], ignore_index=True)
    return df


def _remove_empty_rows(rows: List[Dict]) -> List[Dict]:
    """Remove rows where all values are empty or None."""
    cleaned = []
    for row in rows:
        vals = [str(v).strip() for v in row.values() if v]
        non_empty = [v for v in vals if v and v.lower() not in ("", "none", "nan")]
        if non_empty:
            cleaned.append(row)
    return cleaned


def _remove_empty_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Remove columns where all values (including header) are empty/None."""
    cols_to_keep = []
    for col in df.columns:
        col_name_clean = clean_cell(col)
        vals = df[col].apply(clean_cell).tolist()
        all_vals = [col_name_clean] + vals
        if any(v for v in all_vals):
            cols_to_keep.append(col)
    return df[cols_to_keep]


def _merge_split_headers(raw_table: list) -> list:
    """
    Detect and fix tables where the header is split across 2 rows.
    E.g. Row 0: ['Descripti', 'Tax', 'Adjustment', ...]
         Row 1: ['on',        'payable', 'of negative...', ...]
    Merges them into: ['Description', 'Tax payable', ...]
    """
    if not raw_table or len(raw_table) < 3:
        return raw_table

    row0 = raw_table[0]
    row1 = raw_table[1]

    if len(row0) != len(row1):
        return raw_table

    # Heuristic: if row1 has mostly non-numeric short text that looks like
    # continuations (e.g. 'on', 'payable', 'Payable'), merge as header
    numeric_count = 0
    continuation_count = 0
    for v0, v1 in zip(row0, row1):
        s0 = str(v0).strip() if v0 else ""
        s1 = str(v1).strip() if v1 else ""
        try:
            float(s1.replace(",", ""))
            numeric_count += 1
        except ValueError:
            if s1 and not s1.startswith("("):
                continuation_count += 1

    # If most row1 values look like text continuations (not data numbers),
    # it's likely a split header
    if continuation_count >= len(row0) * 0.4 and numeric_count < len(row0) * 0.3:
        merged_header = []
        for v0, v1 in zip(row0, row1):
            s0 = str(v0).strip() if v0 else ""
            s1 = str(v1).strip() if v1 else ""
            if s0 and s1:
                combined = s0 + " " + s1
            elif s0:
                combined = s0
            else:
                combined = s1
            # Clean newlines
            combined = combined.replace("\n", " ").strip()
            combined = re.sub(r'\s+', ' ', combined)
            merged_header.append(combined)
        return [merged_header] + raw_table[2:]

    return raw_table


def extract_metadata_from_text(text: str) -> Dict[str, str]:
    """Extract GSTIN, Period, Year, Date of ARN from the PDF text."""
    meta = {"gstin": "", "period": "", "year": "", "arn_date": ""}

    # GSTIN pattern — 15-char alphanumeric
    gstin_match = re.search(r'\b(\d{2}[A-Z]{5}\d{4}[A-Z]\d[A-Z\d]{2})\b', text)
    if gstin_match:
        meta["gstin"] = gstin_match.group(1)

    # Months
    months = [
        "January", "February", "March", "April", "May", "June",
        "July", "August", "September", "October", "November", "December",
    ]

    # Period — look for month names near keywords
    period_match = re.search(
        r'(?:Tax\s*Period|Period|Return\s*Period)\s*[:\-]?\s*('
        + '|'.join(months) + r')\s*[\-,]?\s*(\d{4})',
        text, re.IGNORECASE,
    )
    if period_match:
        meta["period"] = period_match.group(1).strip()
        meta["year"] = period_match.group(2).strip()
    else:
        for month in months:
            pat = re.search(rf'\b({month})\s*[\-,]?\s*(\d{{4}})\b', text, re.IGNORECASE)
            if pat:
                meta["period"] = pat.group(1)
                meta["year"] = pat.group(2)
                break

    # Date of ARN
    arn_match = re.search(
        r'(?:ARN|Date\s*of\s*ARN)\s*[:\-]?\s*(\d{2}[/\-]\d{2}[/\-]\d{4})',
        text, re.IGNORECASE,
    )
    if arn_match:
        meta["arn_date"] = arn_match.group(1)
    else:
        date_match = re.search(
            r'(?:Date\s*of\s*[Ff]iling|Filed\s*on)\s*[:\-]?\s*(\d{2}[/\-]\d{2}[/\-]\d{4})',
            text, re.IGNORECASE,
        )
        if date_match:
            meta["arn_date"] = date_match.group(1)

    # Financial year
    fy_match = re.search(r'(\d{4})\s*-\s*(\d{2,4})', text)
    if fy_match:
        fy_start = fy_match.group(1)
        fy_end = fy_match.group(2)
        if len(fy_end) == 2:
            fy_end = fy_start[:2] + fy_end
        meta["year"] = f"{fy_start}-{fy_end[-2:]}"

    return meta


def detect_section(table_text: str, columns: List[str]) -> Optional[str]:
    """Detect which GSTR-3B section a table belongs to."""
    combined = (table_text + " " + " ".join(columns)).lower()

    best_section = None
    best_score = 0

    for section_id, config in GSTR3B_SECTIONS.items():
        score = 0
        for keyword in config["keywords"]:
            if re.search(keyword, combined, re.IGNORECASE):
                score += 1
        if score > best_score:
            best_score = score
            best_section = section_id

    return best_section if best_score > 0 else None


def _extract_tables_pdfplumber(pdf_path: str) -> List[pd.DataFrame]:
    """Extract tables from PDF using pdfplumber."""
    all_dfs = []
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables()
                if tables:
                    for table in tables:
                        if table and len(table) >= 2:
                            # Try to merge split headers (like section 6)
                            table = _merge_split_headers(table)
                            # Use first row as headers
                            headers = []
                            for i, h in enumerate(table[0]):
                                cleaned = str(h).strip() if h else ""
                                if not cleaned or cleaned.lower() in ("none", "nan"):
                                    cleaned = f"Col_{i}"
                                headers.append(cleaned)
                            rows = table[1:]
                            df = pd.DataFrame(rows, columns=headers)
                            all_dfs.append(df)
    except Exception as e:
        logger.error(f"pdfplumber extraction failed: {e}")
    return all_dfs


def _extract_tables_tabula(pdf_path: str) -> List[pd.DataFrame]:
    """Extract tables from PDF using tabula-py."""
    if not HAS_TABULA:
        return []

    all_dfs = []
    try:
        # Try lattice mode first (bordered tables — best for GSTR)
        dfs = tabula.read_pdf(
            pdf_path, pages="all", multiple_tables=True,
            lattice=True, silent=True,
        )
        if dfs and any(not df.empty for df in dfs):
            all_dfs = [df for df in dfs if not df.empty]
        else:
            # Fallback to stream mode
            dfs = tabula.read_pdf(
                pdf_path, pages="all", multiple_tables=True,
                stream=True, silent=True,
            )
            all_dfs = [df for df in dfs if not df.empty] if dfs else []
    except Exception as e:
        logger.warning(f"Tabula failed: {e}")
        return []

    return all_dfs


def extract_gstr3b_tables(pdf_path: str, use_tabula: bool = True) -> Dict[str, Any]:
    """
    Extract all tables from a GSTR-3B PDF, map to sections, attach metadata.
    """
    # 1. Extract full text for metadata
    full_text = ""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            total_pages = len(pdf.pages)
            for page in pdf.pages:
                full_text += (page.extract_text() or "") + "\n"
    except Exception as e:
        logger.error(f"Error reading PDF text: {e}")
        total_pages = 0

    metadata = extract_metadata_from_text(full_text)

    # 2. Extract tables — try tabula first, pdfplumber as fallback
    all_dfs = []
    if use_tabula and HAS_TABULA:
        try:
            all_dfs = _extract_tables_tabula(pdf_path)
        except Exception:
            pass

    if not all_dfs:
        all_dfs = _extract_tables_pdfplumber(pdf_path)

    # 3. Clean and fix each DataFrame
    cleaned_dfs = []
    for df in all_dfs:
        if df.empty or len(df) == 0:
            continue
        # Fix columns that are actually data values
        df = _fix_tabula_headers(df)
        # Remove empty columns
        df = _remove_empty_columns(df)
        if df.empty or len(df.columns) == 0:
            continue
        cleaned_dfs.append(df)

    # 4. Map tables to sections
    sections = {}
    unclassified = []

    for df_idx, df in enumerate(cleaned_dfs):
        # Clean all cells
        for col in df.columns:
            df[col] = df[col].apply(clean_cell)

        # Build text for section detection (columns + all data)
        col_text = " ".join(str(c) for c in df.columns)
        data_text = " ".join(df.astype(str).values.flatten())
        table_text = col_text + " " + data_text[:1000]

        section_id = detect_section(table_text, [str(c) for c in df.columns])

        # Clean column names
        clean_cols = [clean_cell(c) if clean_cell(c) else f"Col_{i}"
                      for i, c in enumerate(df.columns)]

        # Deduplicate column names
        seen = {}
        deduped_cols = []
        for col in clean_cols:
            if col in seen:
                seen[col] += 1
                deduped_cols.append(f"{col}_{seen[col]}")
            else:
                seen[col] = 0
                deduped_cols.append(col)
        clean_cols = deduped_cols

        # Convert to list-of-dicts
        rows = []
        for _, row in df.iterrows():
            row_dict = {}
            for orig_col, clean_col in zip(df.columns, clean_cols):
                val = clean_cell(row[orig_col])
                row_dict[clean_col] = val
            rows.append(row_dict)

        # Remove empty rows
        rows = _remove_empty_rows(rows)

        if not rows:
            continue

        if section_id:
            if section_id in sections:
                # Cross-page fragment: re-map rows to use the first fragment's columns
                existing_cols = sections[section_id]["columns"]
                if len(clean_cols) == len(existing_cols):
                    remapped_rows = []
                    for row in rows:
                        new_row = {}
                        for old_key, new_key in zip(clean_cols, existing_cols):
                            new_row[new_key] = row.get(old_key, "")
                        remapped_rows.append(new_row)
                    sections[section_id]["rows"].extend(remapped_rows)
                else:
                    sections[section_id]["rows"].extend(rows)
            else:
                sections[section_id] = {
                    "title": GSTR3B_SECTIONS[section_id]["title"],
                    "columns": clean_cols,
                    "rows": rows,
                    "row_count": 0,
                }
            sections[section_id]["row_count"] = len(sections[section_id]["rows"])
        else:
            # Auto-name unclassified tables based on content
            all_text = " ".join(" ".join(str(v) for v in r.values()) for r in rows).lower()
            col_text_lower = " ".join(clean_cols).lower()

            if any(k in col_text_lower or k in all_text for k in ["year", "period"]):
                if any(k in all_text for k in ["gstin", "legal name", "trade name", "registered person"]):
                    table_name = "Taxpayer Details"
                else:
                    table_name = "Filing Details"
            elif any(k in col_text_lower or k in all_text for k in ["gstin", "legal name", "trade name"]):
                table_name = "Taxpayer Details"
            elif len(clean_cols) >= 4 and any(k in col_text_lower for k in ["integrated", "central", "state"]):
                # Looks like a breakup / summary row
                table_name = "Breakup"
            else:
                table_name = f"Table {df_idx + 1}"

            # Merge breakup fragments together
            if table_name == "Breakup":
                if "Breakup" not in sections:
                    sections["Breakup"] = {
                        "title": "Breakup - Tax Summary",
                        "columns": clean_cols,
                        "rows": rows,
                        "row_count": 0,
                    }
                else:
                    # Re-map columns if same count
                    existing_cols = sections["Breakup"]["columns"]
                    if len(clean_cols) == len(existing_cols):
                        for row in rows:
                            new_row = {}
                            for old_key, new_key in zip(clean_cols, existing_cols):
                                new_row[new_key] = row.get(old_key, "")
                            sections["Breakup"]["rows"].append(new_row)
                    else:
                        sections["Breakup"]["rows"].extend(rows)
                sections["Breakup"]["row_count"] = len(sections["Breakup"]["rows"])
            else:
                unclassified.append({
                    "name": table_name,
                    "columns": clean_cols,
                    "rows": rows,
                    "row_count": len(rows),
                })

    # 5. Build backward-compat tables list
    tables = []
    for section_id, section in sections.items():
        tables.append({
            "section": section_id,
            "name": section["title"],
            "columns": section["columns"],
            "rows": section["rows"],
            "row_count": section["row_count"],
            "page": 1,
        })
    for ut in unclassified:
        tables.append({
            "section": "other",
            "name": ut["name"],
            "columns": ut["columns"],
            "rows": ut["rows"],
            "row_count": ut["row_count"],
            "page": 1,
        })

    return {
        "metadata": metadata,
        "total_pages": total_pages,
        "sections": sections,
        "unclassified": unclassified,
        "tables": tables,
    }


def build_consolidation_excel(
    extractions: List[Dict[str, Any]],
    output_path: str,
) -> str:
    """
    Build consolidated Excel: one sheet per GSTR-3B section + Breakup summary.
    Rows from all months stacked vertically with Year/Period/ARN/GSTIN metadata.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Collect all section data across all extractions
    merged_sections: Dict[str, List[Dict]] = {}
    breakup_rows = []

    for ext in extractions:
        meta = ext.get("metadata", {})
        year = meta.get("year", "")
        period = meta.get("period", "")
        arn_date = meta.get("arn_date", "")
        gstin = meta.get("gstin", "")

        for section_id, section_data in ext.get("sections", {}).items():
            if section_id not in merged_sections:
                merged_sections[section_id] = []

            for row in section_data.get("rows", []):
                enriched = dict(row)
                enriched["Year"] = year
                enriched["Period"] = period
                enriched["Date of ARN"] = arn_date
                enriched["GSTIN"] = gstin
                merged_sections[section_id].append(enriched)

        # Build breakup row
        breakup_row = {
            "Period": f"{period} {year.split('-')[0] if '-' in year else year}" if period else "",
            "Integrated tax": 0.0,
            "Central tax": 0.0,
            "State/UT tax": 0.0,
            "Cess": 0.0,
            "Date of ARN": arn_date,
            "GSTIN": gstin,
        }
        if "3.1" in ext.get("sections", {}):
            for row in ext["sections"]["3.1"].get("rows", []):
                for k, v in row.items():
                    val = str(v).replace(",", "").replace("-", "0").strip()
                    try:
                        num = float(val)
                    except (ValueError, TypeError):
                        continue
                    kl = k.lower()
                    if "integrated" in kl:
                        breakup_row["Integrated tax"] += num
                    elif "central" in kl:
                        breakup_row["Central tax"] += num
                    elif "state" in kl or "ut" in kl:
                        breakup_row["State/UT tax"] += num
                    elif kl == "cess":
                        breakup_row["Cess"] += num
        breakup_rows.append(breakup_row)

    # Build workbook
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    section_order = ["3.1", "3.1.1", "3.1.2", "4", "5", "5.1", "6"]
    for section_id in section_order:
        rows = merged_sections.get(section_id, [])
        ws = wb.create_sheet(title=section_id)

        if not rows:
            ws["A1"] = f"No data for section {section_id}"
            continue

        columns = list(rows[0].keys())

        # Write headers
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = col_name
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        # Write data
        for row_idx, row_data in enumerate(rows, start=2):
            for col_idx, col_name in enumerate(columns, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = row_data.get(col_name, "")
                cell.border = border

        # Auto-width
        for col_idx in range(1, len(columns) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

    # Breakup sheet
    ws = wb.create_sheet(title="Breakup")
    if breakup_rows:
        columns = list(breakup_rows[0].keys())
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = col_name
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        for row_idx, row_data in enumerate(breakup_rows, start=2):
            for col_idx, col_name in enumerate(columns, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = row_data.get(col_name, "")
                cell.border = border
        for col_idx in range(1, len(columns) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

    wb.save(output_path)
    return output_path
