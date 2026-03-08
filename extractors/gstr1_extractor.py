"""
GSTR-1 Section-Aware Table Extractor

Extracts tables from GSTR-1 PDFs and maps them to their
specific section numbers (4A, 4B, 5, 6A–6C, 7, 8, 9A–9C, 10–15).
Also extracts metadata: GSTIN, Filing Period, ARN, ARN Date.
"""

import re
import logging
import pdfplumber
import pandas as pd
from typing import Dict, Any, List, Optional
from pathlib import Path

logger = logging.getLogger(__name__)

# Reuse clean_cell from gstr3b_extractor
from .gstr3b_extractor import clean_cell

# ──────────────────────────────────────────────────────────────────────
# Section definitions — ordered as they appear in the GSTR-1 PDF
# ──────────────────────────────────────────────────────────────────────
GSTR1_SECTIONS = {
    "4A": {
        "title": "4A - B2B Regular",
        "pattern": r"4A\s*[-–].*B2B\s*Regular",
    },
    "4B": {
        "title": "4B - B2B Reverse Charge",
        "pattern": r"4B\s*[-–].*reverse\s*charge",
    },
    "5": {
        "title": "5 - B2CL (Large)",
        "pattern": r"^5\s*[-–].*B2CL",
    },
    "6A": {
        "title": "6A - Exports",
        "pattern": r"6A\s*[-–—].*[Ee]xport",
    },
    "6B": {
        "title": "6B - SEZ Supplies",
        "pattern": r"6B\s*[-–].*SEZ",
    },
    "6C": {
        "title": "6C - Deemed Exports",
        "pattern": r"6C\s*[-–].*[Dd]eemed",
    },
    "7": {
        "title": "7 - B2CS (Others)",
        "pattern": r"^7\s*[-–].*B2CS",
    },
    "8": {
        "title": "8 - Nil/Exempted/Non-GST",
        "pattern": r"^8\s*[-–].*[Nn]il\s*rated",
    },
    "9A-B2B": {
        "title": "9A - Amendment B2B Regular",
        "pattern": r"9A\s*[-–].*Amendment.*table\s*4\s*[-–].*B2B\s*Regular",
    },
    "9A-B2B-RC": {
        "title": "9A - Amendment B2B Reverse Charge",
        "pattern": r"9A\s*[-–].*Amendment.*table\s*4\s*[-–].*B2B\s*Reverse",
    },
    "9A-B2CL": {
        "title": "9A - Amendment B2CL (Large)",
        "pattern": r"9A\s*[-–].*Amendment.*table\s*5\s*[-–].*B2CL",
    },
    "9A-EXP": {
        "title": "9A - Amendment Exports",
        "pattern": r"9A\s*[-–].*Amendment.*[Ee]xport.*table\s*6A",
    },
    "9A-SEZ": {
        "title": "9A - Amendment SEZ",
        "pattern": r"9A\s*[-–].*Amendment.*SEZ.*table\s*6B",
    },
    "9A-DE": {
        "title": "9A - Amendment Deemed Exports",
        "pattern": r"9A\s*[-–].*Amendment.*[Dd]eemed.*table\s*6C",
    },
    "9B-CDNR": {
        "title": "9B - Credit/Debit Notes (Registered) - CDNR",
        "pattern": r"9B\s*[-–].*Credit.*Debit.*Registered.*CDNR(?!A)",
    },
    "9B-CDNUR": {
        "title": "9B - Credit/Debit Notes (Unregistered) - CDNUR",
        "pattern": r"9B\s*[-–].*Credit.*Debit.*Unregistered.*CDNUR(?!A)",
    },
    "9C-CDNRA": {
        "title": "9C - Amended Credit/Debit Notes (Registered) - CDNRA",
        "pattern": r"9C\s*[-–].*Amended.*Credit.*Debit.*Registered.*CDNRA",
    },
    "9C-CDNURA": {
        "title": "9C - Amended Credit/Debit Notes (Unregistered) - CDNURA",
        "pattern": r"9C\s*[-–].*Amended.*Credit.*Debit.*Unregistered.*CDNURA",
    },
    "10": {
        "title": "10 - Amendment B2C (Others)",
        "pattern": r"^10\s*[-–].*Amendment.*B2C",
    },
    "11A": {
        "title": "11A - Advances Received",
        "pattern": r"11A\(1\).*11A\(2\)\s*[-–].*Advances\s*received",
    },
    "11B": {
        "title": "11B - Advances Adjusted",
        "pattern": r"11B\(1\).*11B\(2\)\s*[-–].*Advance\s*amount\s*received\s*in\s*earlier",
    },
    "11A-Amend": {
        "title": "11A - Amendment to Advances Received",
        "pattern": r"^11A\s*[-–].*Amendment\s*to\s*advances\s*received",
    },
    "11B-Amend": {
        "title": "11B - Amendment to Advances Adjusted",
        "pattern": r"^11B\s*[-–].*Amendment\s*to\s*advances\s*adjusted",
    },
    "12": {
        "title": "12 - HSN-wise Summary",
        "pattern": r"^12\s*[-–].*HSN",
    },
    "13": {
        "title": "13 - Documents Issued",
        "pattern": r"^13\s*[-–].*Documents\s*issued",
    },
    "14": {
        "title": "14 - E-Commerce Supplies",
        "pattern": r"^14\s*[-–].*Supplies\s*made\s*through\s*E-Commerce",
    },
    "14A": {
        "title": "14A - Amended E-Commerce Supplies",
        "pattern": r"^14A\s*[-–].*Amended\s*Supplies\s*made\s*through\s*E-Commerce",
    },
    "15": {
        "title": "15 - Supplies U/s 9(5)",
        "pattern": r"^15\s*[-–].*Supplies\s*U/s\s*9\(5\)",
    },
    "15A-Reg": {
        "title": "15A(I) - Amended Supplies U/s 9(5) - Registered",
        "pattern": r"15A\s*\(I\)\s*[-–].*Amended.*9\(5\).*Registered",
    },
    "15A-Unreg": {
        "title": "15A(II) - Amended Supplies U/s 9(5) - Unregistered",
        "pattern": r"15A\s*\(II\)\s*[-–].*Amended.*9\(5\).*Unregistered",
    },
}

# Standard columns in the GSTR-1 summary table
GSTR1_COLUMNS = [
    "Description",
    "No. of records",
    "Document Type",
    "Value (₹)",
    "Integrated Tax (₹)",
    "Central Tax (₹)",
    "State/UT Tax (₹)",
    "Cess (₹)",
]


def _clean_description(text: str) -> str:
    """Clean a description cell — remove leading watermark letters, newlines."""
    if not text:
        return ""
    # Remove leading single-letter watermark artifacts like "A\n", "L\n", "F\n", "N\n", "I\n"
    text = re.sub(r'^[A-Z]\s*\n\s*', '', text)
    text = text.replace('\n', ' ').replace('\r', '')
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def _is_section_header_row(row: list) -> bool:
    """
    Detect if a row is a section header.
    In GSTR-1 PDFs, section headers have the description in col 0
    and all other cols are None/empty.
    """
    if not row or len(row) < 2:
        return False
    desc = str(row[0]).strip() if row[0] else ""
    if not desc or desc.lower() in ("", "none", "nan"):
        return False
    # Check if remaining cells are all empty/None
    rest = row[1:]
    for cell in rest:
        s = str(cell).strip() if cell else ""
        if s and s.lower() not in ("", "none", "nan"):
            return False
    return True


def _is_empty_row(row: list) -> bool:
    """Check if all cells in a row are empty/None."""
    for cell in row:
        s = str(cell).strip() if cell else ""
        if s and s.lower() not in ("", "none", "nan"):
            return False
    return True


def _is_header_row(row: list) -> bool:
    """Detect the repeating column header row (Description, No. of records, ...)"""
    if not row or len(row) < 4:
        return False
    desc = str(row[0]).strip().lower() if row[0] else ""
    return desc == "description"


def _detect_section_id(description: str) -> Optional[str]:
    """Match a section header description to a section ID."""
    cleaned = _clean_description(description)
    for section_id, config in GSTR1_SECTIONS.items():
        if re.search(config["pattern"], cleaned, re.IGNORECASE):
            return section_id
    return None


def _detect_special_rows(description: str) -> Optional[str]:
    """Detect special rows like Total Liability."""
    cleaned = _clean_description(description)
    if re.search(r"Total\s*Liability.*Outward\s*supplies", cleaned, re.IGNORECASE):
        return "Total Liability"
    return None


def extract_gstr1_metadata(pdf_path: str) -> Dict[str, str]:
    """
    Extract metadata from the first page of a GSTR-1 PDF:
    GSTIN, Legal Name, Trade Name, ARN, ARN Date, Financial Year, Tax Period.
    """
    meta = {
        "gstin": "",
        "legal_name": "",
        "trade_name": "",
        "arn": "",
        "arn_date": "",
        "year": "",
        "period": "",
    }

    try:
        with pdfplumber.open(pdf_path) as pdf:
            if not pdf.pages:
                return meta
            page = pdf.pages[0]
            tables = page.extract_tables()

            for table in tables:
                if not table:
                    continue
                for row in table:
                    if not row or len(row) < 2:
                        continue

                    # 2-col metadata table: Financial year / Tax period
                    if len(row) == 2:
                        key = str(row[0]).strip().lower() if row[0] else ""
                        val = str(row[1]).strip() if row[1] else ""
                        if "financial year" in key:
                            meta["year"] = val
                        elif "tax period" in key:
                            meta["period"] = val

                    # 4-col detail table: GSTIN, Legal Name, ARN etc.
                    if len(row) >= 4:
                        # Check col index 1 or 2 for field labels
                        label = ""
                        value = ""
                        if row[1] and row[2]:
                            label = str(row[2]).strip().lower() if row[2] else ""
                            value = str(row[3]).strip() if row[3] else ""
                        elif row[1]:
                            label = str(row[1]).strip().lower() if row[1] else ""
                            value = str(row[3]).strip() if len(row) > 3 and row[3] else ""

                        if "gstin" in label:
                            meta["gstin"] = value
                        elif "legal name" in label:
                            meta["legal_name"] = value
                        elif "trade name" in label:
                            meta["trade_name"] = value
                        elif label == "arn" or label == "(c)":
                            # ARN is at (c)
                            if row[2] and "arn" in str(row[2]).lower():
                                meta["arn"] = str(row[3]).strip() if len(row) > 3 and row[3] else ""
                        elif "arn date" in label:
                            meta["arn_date"] = value

            # Fallback: try text extraction for metadata
            text = page.extract_text() or ""

            if not meta["gstin"]:
                gstin_match = re.search(r'\b(\d{2}[A-Z]{5}\d{4}[A-Z]\d[A-Z\d]{2})\b', text)
                if gstin_match:
                    meta["gstin"] = gstin_match.group(1)

            if not meta["arn_date"]:
                arn_date_match = re.search(r'ARN\s*date\s*(\d{2}/\d{2}/\d{4})', text, re.IGNORECASE)
                if arn_date_match:
                    meta["arn_date"] = arn_date_match.group(1)

            if not meta["arn"]:
                arn_match = re.search(r'\(c\)\s*ARN\s+([A-Z0-9]+)', text, re.IGNORECASE)
                if arn_match:
                    meta["arn"] = arn_match.group(1)

    except Exception as e:
        logger.error(f"Error extracting GSTR-1 metadata: {e}")

    return meta


def extract_gstr1_tables(pdf_path: str, use_tabula: bool = False) -> Dict[str, Any]:
    """
    Extract all tables from a GSTR-1 PDF, map to sections, attach metadata.

    The GSTR-1 summary PDF has one continuous table spanning multiple pages.
    Section headers are embedded as rows where only the Description column has text.
    Data rows have values across all 8 columns.

    Returns the same shape as extract_gstr3b_tables for compatibility:
    {metadata, total_pages, sections, unclassified, tables}
    """
    # 1. Extract metadata from page 1
    metadata = extract_gstr1_metadata(pdf_path)

    # 2. Extract all summary table rows across all pages using pdfplumber
    all_data_rows = []  # (section_id or None, row_dict)
    total_pages = 0
    current_section = None

    try:
        with pdfplumber.open(pdf_path) as pdf:
            total_pages = len(pdf.pages)

            for page_idx, page in enumerate(pdf.pages):
                tables = page.extract_tables()
                if not tables:
                    continue

                for table in tables:
                    if not table:
                        continue

                    # Skip the small metadata tables on page 1 (2-col or 4-col with GSTIN etc.)
                    if page_idx == 0 and table:
                        # Check if this is a metadata table (2-col)
                        if len(table[0]) <= 4:
                            # Check if it contains metadata keywords
                            flat_text = " ".join(
                                str(cell).lower() for row in table for cell in row if cell
                            )
                            if any(k in flat_text for k in ["financial year", "tax period", "gstin", "arn"]):
                                continue

                    # Process each row of the summary table
                    for row in table:
                        if not row:
                            continue

                        # Skip the repeating column header row
                        if _is_header_row(row):
                            continue

                        # Skip empty rows
                        if _is_empty_row(row):
                            continue

                        # Check if this is a section header
                        if _is_section_header_row(row):
                            desc = _clean_description(str(row[0]) if row[0] else "")
                            new_section = _detect_section_id(desc)
                            if new_section:
                                current_section = new_section
                            # Also check for special rows like "Unregistered Type"
                            # These are sub-headers within a section, skip them
                            continue

                        # This is a data row — map it to current section
                        if len(row) >= 8:
                            row_dict = {}
                            for col_idx, col_name in enumerate(GSTR1_COLUMNS):
                                cell_val = clean_cell(row[col_idx]) if col_idx < len(row) else ""
                                row_dict[col_name] = cell_val
                            all_data_rows.append((current_section, row_dict))
                        elif len(row) >= 4:
                            # Some rows may have fewer columns (e.g., section 8 Nil/Exempt)
                            row_dict = {}
                            for col_idx in range(len(row)):
                                if col_idx < len(GSTR1_COLUMNS):
                                    cell_val = clean_cell(row[col_idx]) if row[col_idx] else ""
                                    row_dict[GSTR1_COLUMNS[col_idx]] = cell_val
                            # Fill missing columns
                            for col_name in GSTR1_COLUMNS:
                                if col_name not in row_dict:
                                    row_dict[col_name] = ""
                            all_data_rows.append((current_section, row_dict))

    except Exception as e:
        logger.error(f"Error extracting GSTR-1 tables: {e}")

    # 3. Group rows by section
    sections = {}
    unclassified = []
    total_liability_row = None

    for section_id, row_dict in all_data_rows:
        # Check for Total Liability row
        special = _detect_special_rows(row_dict.get("Description", ""))
        if special == "Total Liability":
            total_liability_row = row_dict
            continue

        if section_id and section_id in GSTR1_SECTIONS:
            if section_id not in sections:
                sections[section_id] = {
                    "title": GSTR1_SECTIONS[section_id]["title"],
                    "columns": GSTR1_COLUMNS,
                    "rows": [],
                    "row_count": 0,
                }
            sections[section_id]["rows"].append(row_dict)
            sections[section_id]["row_count"] = len(sections[section_id]["rows"])
        else:
            unclassified.append({
                "name": f"Unclassified",
                "columns": GSTR1_COLUMNS,
                "rows": [row_dict],
                "row_count": 1,
            })

    # Add Total Liability as its own section
    if total_liability_row:
        sections["Total"] = {
            "title": "Total Liability (Outward Supplies)",
            "columns": GSTR1_COLUMNS,
            "rows": [total_liability_row],
            "row_count": 1,
        }

    # 4. Build backward-compat tables list
    tables = []
    for section_id, section in sections.items():
        tables.append({
            "section": section_id,
            "name": section["title"],
            "columns": section["columns"],
            "rows": section["rows"],
            "row_count": section["row_count"],
            "page": 1,
        })
    for ut in unclassified:
        tables.append({
            "section": "other",
            "name": ut["name"],
            "columns": ut["columns"],
            "rows": ut["rows"],
            "row_count": ut["row_count"],
            "page": 1,
        })

    return {
        "metadata": metadata,
        "total_pages": total_pages,
        "sections": sections,
        "unclassified": unclassified,
        "tables": tables,
    }


def build_gstr1_consolidation_excel(
    extractions: List[Dict[str, Any]],
    output_path: str,
) -> str:
    """
    Build consolidated Excel: one sheet per GSTR-1 section + Summary sheet.
    Rows from all months stacked vertically with Year/Period/ARN/GSTIN metadata.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Collect all section data across all extractions
    merged_sections: Dict[str, List[Dict]] = {}
    summary_rows = []

    for ext in extractions:
        meta = ext.get("metadata", {})
        year = meta.get("year", "")
        period = meta.get("period", "")
        arn_date = meta.get("arn_date", "")
        gstin = meta.get("gstin", "")

        for section_id, section_data in ext.get("sections", {}).items():
            if section_id not in merged_sections:
                merged_sections[section_id] = []

            for row in section_data.get("rows", []):
                enriched = dict(row)
                enriched["Year"] = year
                enriched["Period"] = period
                enriched["Date of ARN"] = arn_date
                enriched["GSTIN"] = gstin
                merged_sections[section_id].append(enriched)

        # Build summary row from key sections
        summary_row = {
            "Period": f"{period} {year}" if period else "",
            "B2B Value": "",
            "B2B IGST": "",
            "B2B CGST": "",
            "B2B SGST": "",
            "SEZ Value": "",
            "B2CS Value": "",
            "HSN Total Value": "",
            "Total Liability Value": "",
            "GSTIN": gstin,
            "Date of ARN": arn_date,
        }

        # Extract key totals from sections
        for section_id, section_data in ext.get("sections", {}).items():
            for row in section_data.get("rows", []):
                desc = row.get("Description", "").strip().lower()
                if desc == "total":
                    val = row.get("Value (₹)", "")
                    igst = row.get("Integrated Tax (₹)", "")
                    cgst = row.get("Central Tax (₹)", "")
                    sgst = row.get("State/UT Tax (₹)", "")
                    if section_id == "4A":
                        summary_row["B2B Value"] = val
                        summary_row["B2B IGST"] = igst
                        summary_row["B2B CGST"] = cgst
                        summary_row["B2B SGST"] = sgst
                    elif section_id == "6B":
                        summary_row["SEZ Value"] = val
                    elif section_id == "7":
                        summary_row["B2CS Value"] = val
                    elif section_id == "12":
                        summary_row["HSN Total Value"] = val

            if section_id == "Total":
                for row in section_data.get("rows", []):
                    summary_row["Total Liability Value"] = row.get("Value (₹)", "")

        summary_rows.append(summary_row)

    # Build workbook
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Determine section order — put main sections first, amendments later
    main_order = ["4A", "4B", "5", "6A", "6B", "6C", "7", "8", "12", "13", "14", "15"]
    amend_order = [
        "9A-B2B", "9A-B2B-RC", "9A-B2CL", "9A-EXP", "9A-SEZ", "9A-DE",
        "9B-CDNR", "9B-CDNUR", "9C-CDNRA", "9C-CDNURA",
        "10", "11A", "11B", "11A-Amend", "11B-Amend", "14A",
        "15A-Reg", "15A-Unreg", "Total",
    ]
    section_order = main_order + amend_order

    for section_id in section_order:
        rows = merged_sections.get(section_id, [])
        if not rows:
            continue

        # Truncate sheet name to 31 chars (Excel limit)
        sheet_title = GSTR1_SECTIONS.get(section_id, {}).get("title", section_id)
        if section_id == "Total":
            sheet_title = "Total Liability"
        sheet_title = re.sub(r'[\\/*?:\[\]]', '_', sheet_title)
        sheet_name = sheet_title[:31]

        ws = wb.create_sheet(title=sheet_name)
        columns = list(rows[0].keys())

        # Write headers
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = col_name
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        # Write data
        for row_idx, row_data in enumerate(rows, start=2):
            for col_idx, col_name in enumerate(columns, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = row_data.get(col_name, "")
                cell.border = border

        # Auto-width
        for col_idx in range(1, len(columns) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

    # Summary sheet
    ws = wb.create_sheet(title="Summary", index=0)
    if summary_rows:
        columns = list(summary_rows[0].keys())
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = col_name
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        for row_idx, row_data in enumerate(summary_rows, start=2):
            for col_idx, col_name in enumerate(columns, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = row_data.get(col_name, "")
                cell.border = border
        for col_idx in range(1, len(columns) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

    wb.save(output_path)
    return output_path
