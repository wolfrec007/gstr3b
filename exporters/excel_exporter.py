from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import Dict, Any
import logging

logger = logging.getLogger(__name__)

class ExcelExporter:
    """Export extraction data to Excel format"""
    
    def __init__(self):
        self.header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def export(self, extraction_data: Dict[str, Any], output_path: Path) -> Path:
        """Export extraction to Excel file"""
        try:
            wb = Workbook()
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Create a sheet for each table
            for idx, table in enumerate(extraction_data.get('tables', [])):
                sheet_name = f"{table['name'][:25]}" if len(table['name']) > 25 else table['name']
                
                # Make sheet name unique if duplicate
                if sheet_name in wb.sheetnames:
                    sheet_name = f"{sheet_name}_{idx+1}"
                
                ws = wb.create_sheet(title=sheet_name)
                
                # Write headers
                columns = table.get('columns', [])
                for col_idx, column in enumerate(columns, start=1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.value = column
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = self.border
                
                # Write data rows
                for row_idx, row_data in enumerate(table.get('rows', []), start=2):
                    for col_idx, column in enumerate(columns, start=1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.value = row_data.get(column, "")
                        cell.border = self.border
                        
                        # Right-align numeric columns
                        if self._is_numeric_column(column):
                            cell.alignment = Alignment(horizontal='right')
                
                # Auto-adjust column widths
                for col_idx in range(1, len(columns) + 1):
                    ws.column_dimensions[get_column_letter(col_idx)].width = 15
            
            # If no tables, create a summary sheet
            if not extraction_data.get('tables'):
                ws = wb.create_sheet(title="Summary")
                ws['A1'] = "No tables found in the document"
            
            # Save workbook
            wb.save(output_path)
            logger.info(f"Excel file exported to {output_path}")
            return output_path
        
        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}")
            raise
    
    def export_consolidated(self, consolidated_data: Dict[str, Any], output_path: Path) -> Path:
        """Export consolidated yearly data to Excel"""
        try:
            wb = Workbook()
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Create summary sheet
            summary_ws = wb.create_sheet(title="Summary")
            summary_ws['A1'] = "Yearly Consolidated Report"
            summary_ws['A1'].font = Font(bold=True, size=14)
            summary_ws['A3'] = "Year:"
            summary_ws['B3'] = consolidated_data.get('year', '')
            summary_ws['A4'] = "Form Type:"
            summary_ws['B4'] = consolidated_data.get('form_type', '')
            summary_ws['A5'] = "Total Records:"
            summary_ws['B5'] = consolidated_data.get('total_records', 0)
            
            # Create a sheet for each month
            monthly_data = consolidated_data.get('monthly_data', {})
            for month, records in monthly_data.items():
                ws = wb.create_sheet(title=month[:31])  # Excel sheet name limit
                
                if not records:
                    ws['A1'] = f"No data for {month}"
                    continue
                
                # Get columns from first record
                if records and isinstance(records[0], dict):
                    columns = list(records[0].keys())
                    
                    # Write headers
                    for col_idx, column in enumerate(columns, start=1):
                        cell = ws.cell(row=1, column=col_idx)
                        cell.value = column
                        cell.font = self.header_font
                        cell.fill = self.header_fill
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = self.border
                    
                    # Write data
                    for row_idx, record in enumerate(records, start=2):
                        for col_idx, column in enumerate(columns, start=1):
                            cell = ws.cell(row=row_idx, column=col_idx)
                            cell.value = record.get(column, "")
                            cell.border = self.border
                            
                            if self._is_numeric_column(column):
                                cell.alignment = Alignment(horizontal='right')
                    
                    # Auto-adjust column widths
                    for col_idx in range(1, len(columns) + 1):
                        ws.column_dimensions[get_column_letter(col_idx)].width = 15
            
            wb.save(output_path)
            logger.info(f"Consolidated Excel file exported to {output_path}")
            return output_path
        
        except Exception as e:
            logger.error(f"Error exporting consolidated data to Excel: {e}")
            raise
    
    def _is_numeric_column(self, column_name: str) -> bool:
        """Check if column should be right-aligned (numeric data)"""
        numeric_keywords = ['value', 'amount', 'tax', 'igst', 'cgst', 'sgst', 'cess', 'rate', 'total']
        return any(keyword in column_name.lower() for keyword in numeric_keywords)
